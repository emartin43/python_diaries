
import pandas as pd
import os 
import openpyxl
import datetime

from openpyxl import load_workbook

##SAVE .XLS to .XLSX

import os
import pandas as pd

# Specify the parent folder containing subfolders with .xls files
parent_folder_path = r'\\fs109\es-comm\Commercial_Share_Folder\Bright Rebates\BPA Lighting Calculator Extraction Tool - Rosetta\ExcelLightCalcInbox\Pre 2019'

# Specify the destination folder path for .xlsx files
destination_folder_path = r'\\fs109\ES-Comm\Commercial_Share_Folder\Bright Rebates\BPA Lighting Calculator Extraction Tool - Rosetta\output_file'

# Iterate over subfolders in the parent folder
for subfolder_name in os.listdir(parent_folder_path):
    subfolder_path = os.path.join(parent_folder_path, subfolder_name)

    # Check if the subfolder is a directory
    if os.path.isdir(subfolder_path):
        # Get a list of all files in the subfolder
        file_list = os.listdir(subfolder_path)

        # Loop through each file in the subfolder
        for file_name in file_list:
            # Check if the file is a .xls file
            if file_name.endswith('.xls'):
                xls_path = os.path.join(subfolder_path, file_name)

                try:
                    # Read specific sheets from the .xls file 
                    df = pd.read_excel(xls_path, sheet_name=["user_Lighting Details", "UploadData"])
                    
                    # Concatenate the DataFrames in the dictionary into a single DataFrame
                    combined_df = pd.concat(df.values(), ignore_index=True)
                    
                    # Define the new file name with "loop_" prefix and .xlsx extension
                    new_file_name = "loop_" + os.path.splitext(file_name)[0] + ".xlsx"

                    # Save the DataFrame as .xlsx with the new file name in the destination folder
                    new_xlsx_path = os.path.join(destination_folder_path, new_file_name)
                    combined_df.to_excel(new_xlsx_path, index=False)

                    print(f"Converted '{file_name}' to '{new_file_name}' and saved to destination folder")
                except Exception as e:
                    print(f"Error processing '{file_name}': {str(e)}")
##DATA MANIPULATION TO HAVE ALL INFO IN ONE SHEET

import os
import openpyxl

# Specify the folder containing the Excel files
folder_path = r'\\fs109\ES-Comm\Commercial_Share_Folder\Bright Rebates\BPA Lighting Calculator Extraction Tool - Rosetta\output_file'

# List the Excel files in the folder
excel_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]

# Iterate through each Excel file in the folder
for excel_file in excel_files:
    # Open the workbook
    workbook = openpyxl.load_workbook(os.path.join(folder_path, excel_file))

    # Choose the sheet 
    sheet = workbook['Sheet1']

    # Iterate through the cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell contains 'Retrofit'
            if cell.value == 'Retrofit':
                # Set the initial source cell to one column to the left of the cell containing 'Retrofit'
                source_cell = sheet.cell(row=cell.row, column=cell.column - 1)

                # Set the initial destination cell to Z9
                destination_cell = sheet['Z9']

                # Iterate through 28 columns to the right
                for col_index in range(1, 29):
                    # Move the value from the source cell to the destination cell
                    destination_cell.value = source_cell.value

                    # Clear the source cell 
                    source_cell.value = None

                    # Move both the source and destination cells right one column
                    source_cell = sheet.cell(row=source_cell.row, column=source_cell.column + 1)
                    destination_cell = sheet.cell(row=destination_cell.row, column=destination_cell.column + 1)

                # Delete the cell that contained 'Retrofit'
                sheet.delete_rows(cell.row)

    # Save the changes to the Excel file
    workbook.save(os.path.join(folder_path, excel_file))

    # Close the workbook
    workbook.close()


##Here defining function that scrapes excel spreadsheets and saves them individually


import openpyxl
import pandas as pd
import os

def process_and_combine_data(source_file_path, output_file_path):
    sheet_name = "Sheet1"

    source_workbook = openpyxl.load_workbook(source_file_path, data_only=True)
    source_sheet = source_workbook[sheet_name]

    # existing code for extracting cell_values and cell_names here...
    
    # Extracted cell values
    cell_values = [
        source_sheet['P2'].value,
        source_sheet['P4'].value, source_sheet['P5'].value, source_sheet['P6'].value, source_sheet['P7'].value,
        source_sheet['P3'].value,
        source_sheet['U4'].value,
        source_sheet['U5'].value,
        source_sheet['P10'].value,
        source_sheet['P11'].value,
        source_sheet['P12'].value,
        source_sheet['S10'].value,
        source_sheet['S11'].value,
        source_sheet['S12'].value,
        source_sheet['S13'].value,
        source_sheet['V10'].value.split(":")[1].strip() if source_sheet['V10'].value else " ",  # Check if V10 is empty
        source_sheet['V11'].value.split(":")[1].strip() if source_sheet['V11'].value else " ",
        source_sheet['V12'].value.split(":")[1].strip() if source_sheet['V12'].value else " ",
        source_sheet['Q15'].value, source_sheet['Q16'].value, source_sheet['Q17'].value, source_sheet['Q18'].value,
        source_sheet['Q19'].value,source_sheet['AE9'].value, source_sheet['AF9'].value, source_sheet['AG9'].value, source_sheet['AH9'].value, 
        source_sheet['AI9'].value, source_sheet['AJ9'].value, source_sheet['AL9'].value, source_sheet['AS9'].value,
        source_sheet['AU9'].value,
        source_sheet['AV9'].value, source_sheet['AW9'].value, source_sheet['BA9'].value, source_sheet['BB9'].value, source_sheet['BC9'].value
    ]


    # Define cell names for the columns in the output file
    cell_names = [
        'project_name', 'mailing_address', 'mailing_city_state_zip', 'contact_phone', 'contact_email',
        'company_name', 'project/site_address', 'project/site_city_state_zip', 'tpu_contact_name', 'tpu_contact_phone', 'tpu_email', 'ta_organization',
        'ta_contact_name', 'ta_contact_phone', 'ta_contact_email',
        'class', 'energy_rate_($/kWh)', 'energy_rate_($/kW)', 'est_annual_energy_savings', 'est_energy_savings_(%)', 'est_annual_utility_bill_savings', 'est_total_proj_install_costs', 'est_total_proj_incentive',
        'projcost', 'busbr_sav', 'wattred', 'totcred', 'bc_ratio', 'bldg_type', 'unique_site_id', 'fedblg', 'completion_date', 'bpa_max_wtp', 'funding_source',
        'nwtan_member', 'neea_nxt_lvl_1', 'neea_nxt_level_2'
    ]


    # Fill with empty strings to match the length of cell_names
    cell_values += [''] * (len(cell_names) - len(cell_values))
    
    
    # Extract "mailing_address" and "mailing_city_state_zip" values
    mailing_address = source_sheet['F4'].value
    mailing_city_state_zip = source_sheet['F5'].value
    
    # Extract "project/site_address" and "project/site_city_state_zip" values
    project_site_address = source_sheet['G4'].value
    project_site_city_state_zip = source_sheet['G5'].value
    

    # Create a DataFrame with extracted cell values and column names
    df_existing = pd.DataFrame([cell_values], columns=cell_names)
    
    
     # Define a function to combine "mailing_address" and "mailing_city_state_zip" with a separator
    def combine_address(row):
        return f"{row['mailing_address']} {row['mailing_city_state_zip']}"

    # Apply the function to create a new column "mailing_full_address"
    df_existing['mailing_full_address'] = df_existing.apply(combine_address, axis=1)
    
    # Define a function to combine "project/site_address" and "project/site_city_state_zip" with a separator
    def combine_site_address(row):
        return f"{row['project/site_address']} {row['project/site_city_state_zip']}"

    # Apply the function to create a new column "project/site_full_address"
    df_existing['project/site_full_address'] = df_existing.apply(combine_site_address, axis=1)
    


    #existing code for extracting and merging additional data here...
    
    # Create an empty DataFrame to store the extracted values
    columns_to_extract = ['K', 'L', 'M', 'N', 'P', 'Q', 'R', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD'
                         , 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 
                          'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 
                          'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 
                          'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 
                          'CO', 'CP', 'CQ', 'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ', 'DA', 'DB', 'DC', 'DD', 
                          'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 
                          'DU', 'DV', 'DW', 'DX', 'DY', 'DZ', 'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EH', 'EI', 'EJ', 
                          'EK', 'EL', 'EM', 'EN', 'EO', 'EP', 'EQ', 'ER', 'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ', 
                          'FA', 'FB', 'FC', 'FD', 'FE', 'FF'
]
    df_existing2 = pd.DataFrame(columns=columns_to_extract + [''])

    # Initialize a variable to keep track of the row index
    row_index = 22  # Start from row 22 (assuming 1-based index)

    # Get the value from cell P2 of the source sheet
    project_name_value = source_sheet['P2'].value
    
    


    # Loop through each row in the source sheet
    while True:
        # Extract values from columns K to FF for the current row
        values = [
            source_sheet[f'{col}{row_index}'].value
            for col in columns_to_extract
        ]
    
        # Check if all values in the extracted row are None (blank)
        if all(value is None for value in values):
            break  # Break the loop if an entirely blank row is found
    
        #Add the value from project_name and the extracted values to df_existing2
        df_existing2.loc[len(df_existing2)] = values + ['']
    
        # Move to the next row
        row_index += 1

    # Create a new DataFrame by repeating df_existing based on the length of df_existing2
    repeated_df_existing = pd.concat([df_existing] * len(df_existing2), ignore_index=True)

    # Concatenate the two DataFrames side by side
    df_combined = pd.concat([repeated_df_existing, df_existing2], axis=1)

    # existing code for renaming columns 
    
    # Define the path to the folder containing the files
    folder_path = r'\\fs109\es-comm\Commercial_Share_Folder\Bright Rebates\BPA Lighting Calculator Extraction Tool - Rosetta\ExcelLightCalcInbox\Pre 2019'  
    
    
    
    # Rename columns in df_existing
    column_rename_dict = {
        'K': 'measure_no', 'L': 'status', 'M': 'space', 'N': 'annual_hrs',
        'P': 'existing', 'Q': 'proposed', 'R': 'controls', 'T': 'kWh/year',
        'U': 'watts_per_fixture', 'V': 'incentives', 'W': 'notes', 'X': 'blank_column', 'Y': 'kw_for_table', 'Z': 'space', 
        'AA': 'existing_quantity', 'AB': 'existing_class', 'AC': 'existing_type', 'AD': 'existing_subtype', 
        'AE': 'existing_lamp_wattage', 'AF': 'existing_lamps_per_fixture', 'AG': 'proposed_measure_type', 
        'AH': 'proposed_quantity', 'AI': 'proposed_class', 'AJ': 'proposed_type', 'AK': 'proposed_subtype', 
        'AL': 'proposed_lamp_wattage', 'AM': 'proposed_lamps_per_fixture', 'AN': 'proposed_lamp_model_number', 
        'AO': 'proposed_ballast_model_number', 'AP': 'no_controls', 'AQ': 'controls_quantity', 'AR': 'controls_class', 
        'AS': 'controls_percent_reduction', 'AT': 'notes', 'AU': 'number_of_errors', 'AV': 'existing_mml_row', 
        'AW': 'proposed_mml_row', 'AX': 'existing_ballast_factor', 'AY': 'proposed_ballast_factor', 
        'AZ': 'existing_fixture_wattage', 'BA': 'proposed_fixture_wattage', 'BB': 'arith_fixture_wattage', 
        'BC': 'electric_interaction', 'BD': 'gas_interaction', 'BE': 'existing_fixture_kwh_base', 
        'BF': 'arith_existing_fixture_kwh_base', 'BG': 'proposed_fixture_kwh_base', 'BH': 'decommissioned_fixtures', 
        'BI': 'precond_decom_kwh_savings', 'BJ': 'arith_decom_kwh_savings', 'BK': 'precond_equip_kwh_savings', 
        'BL': 'arith_equip_kwh_savings', 'BM': 'precond_controls_kwh_savings', 'BN': 'arith_controls_kwh_savings', 
        'BO': 'precond_decom_and_equip_kwh_savings', 'BP': 'arith_decom_and_equip_kwh_savings', 
        'BQ': 'precond_total_kwh_savings', 'BR': 'arith_total_kwh_savings', 'BS': 'precond_equip_and_controls_kwh_savings',
        'BT': 'arith_equip_and_controls_kwh_savings', 'BU': 'precond_equip_and_controls_percent_savings', 
        'BV': 'arith_equip_and_controls_percent_savings', 'BW': 'precond_measure_percent_savings', 
        'BX': 'arith_measure_percent_savings', 'BY': 'precond_measure_kw', 'BZ': 'arith_measure_kw', 
        'CA': 'proposed_measure_kw', 'CB': 'precond_kw_savings', 'CC': 'arith_kw_savings', 
        'CD': 'existing_fixture_kwh_with_hvac', 'CE': 'arith_existing_fixture_kwh_with_hvac', 
        'CF': 'proposed_fixture_kwh_with_hvac', 'CG': 'precond_decom_kwh_savings_with_hvac', 
        'CH': 'arith_decom_kwh_savings_with_hvac', 'CI': 'precond_equip_kwh_savings_with_hvac', 
        'CJ': 'arith_equip_kwh_savings_with_hvac', 'CK': 'precond_equip_and_controls_kwh_savings_with_hvac', 
        'CL': 'arith_equip_and_controls_kwh_savings_with_hvac', 'CM': 'precond_controls_kwh_savings_with_hvac', 
        'CN': 'arith_controls_kwh_savings_with_hvac', 'CO': 'precond_decom_and_equip_kwh_savings_with_hvac',
        'CP': 'arith_decom_and_equip_kwh_savings_with_hvac', 'CQ': 'precond_total_kwh_savings_with_hvac', 
        'CR': 'arith_total_kwh_savings_with_hvac', 'CS': 'existing_fixture_kwh_with_hvac_busbar', 
        'CT': 'arith_existing_fixture_kwh_with_hvac_busbar', 'CU': 'proposed_fixture_kwh_with_hvac_busbar', 
        'CV': 'precond_decom_kwh_savings_with_hvac_busbar', 'CW': 'arith_decom_kwh_savings_with_hvac_busbar', 'CX': 'precond_equip_kwh_savings_with_hvac_busbar', 'CY': 'arith_equip_kwh_savings_with_hvac_busbar', 'CZ': 'precond_equip_and_controls_kwh_savings_with_hvac_busbar', 'DA': 'arith_equip_and_controls_kwh_savings_with_hvac_busbar', 'DB': 'precond_controls_kwh_savings_with_hvac_busbar', 'DC': 'arith_controls_kwh_savings_with_hvac_busbar', 'DD': 'precond_decom_and_equip_kwh_savings_with_hvac_busbar', 'DE': 'arith_decom_and_equip_kwh_savings_with_hvac_busbar', 'DF': 'precond_total_kwh_savings_with_hvac_busbar', 'DG': 'arith_total_kwh_savings_with_hvac_busbar', 'DH': 'precond_total_measure_kwh_with_hvac_busbar', 'DI': 'arith_total_measure_kwh_with_hvac_busbar', 'DJ': 'proposed_total_measure_kwh_with_hvac_busbar', 'DK': 'therms_increase_precond', 'DL': 'therms_increase_arith', 'DM': 'decom_incentive_rates', 'DN': 'utility_decom_incentive_rate', 'DO': 'bpa_decom_incentive_rate', 'DP': 'nonstandard_incentive_rates', 'DQ': 'utility_nonstandard_incentive_rate', 'DR': 'bpa_nonstandard_incentive_rate', 'DS': 'controls_incentive_quantity', 'DT': 'controls_incentive_utility', 'DU': 'controls_incentive_bpa', 'DV': 'controls_text', 'DW': 'controls_incentive_vector', 'DX': 'controls_incentive', 'DY': 'bpa_controls_incentive_rate', 'DZ': 'utility_controls_incentive_rate', 'EA': 'decom_incentive_quantity', 'EB': 'decom_incentive_utility', 'EC': 'decom_incentive_bpa', 'ED': 'equipment_incentive_quantity', 'EE': 'equipment_incentive_utility', 'EF': 'equipment_incentive_bpa', 'EG': 'measure_text', 'EH': 'decommissioning_and_equipment_results_text', 'EI': 'controls_results_text', 'EJ': 'precond_total_measure_kwh', 'EK': 'arith_total_measure_kwh', 'EL': 'proposed_total_measure_kwh', 'EM': 'precond_total_measure_kwh_with_hvac', 'EN': 'arith_total_measure_kwh_with_hvac', 'EO': 'proposed_total_measure_kwh_with_hvac', 'EP': 'precond_kw_percent_savings', 'EQ': 'arith_kw_percent_savings', 'ER': 'total_utility_incentive', 'ES': 'total_bpa_incentive', 'ET': 'non_standard_status', 'EU': 'status_notes', 'EV': 'control_kwh_savings', 'EW': 'note_that_this_is_actually_a_percent_savings_0_to_100', 'EX': 'utility_equipment_incentive_rate', 'EY': 'decom_and_equipment_incentive_name', 'EZ': 'decom_and_equipment_incentive_description', 'FA': 'controls_incentive_name', 'FB': 'controls_incentive_description', 'FC': 'approved_non_standard_user_field_concat', 'FD': 'space_type', 'FE': 'ref_no', 'FF': 'equals_percent_of_time_lights_are_turned_off_by_controls'


    }
    df_combined.rename(columns=column_rename_dict, inplace=True)


    # Save the merged DataFrame to the output file
    df_combined.to_excel(output_file_path, index=False)

    print(f"Data from '{source_file_path}' added to the output file '{output_file_path}'.")



# Specify the folder containing the Excel files
folder_path = r'\\fs109\ES-Comm\Commercial_Share_Folder\Bright Rebates\BPA Lighting Calculator Extraction Tool - Rosetta\output_file'

# Define the folder path for the output files 
output_folder = r'\\fs109\ES-Comm\Commercial_Share_Folder\Bright Rebates\BPA Lighting Calculator Extraction Tool - Rosetta\output_loop_file'



# Get a list of all files in the folder
file_list = os.listdir(folder_path)

for file_name in file_list:
    # Check if the file is an Excel file (assuming all files in the folder are Excel files)
    if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
        source_file_path = os.path.join(folder_path, file_name)
        
        # Adjust the output file name format, using the original file name without extension as a prefix
        output_file_name = os.path.splitext(file_name)[0] + '_output.xlsx'
        output_file_path = os.path.join(output_folder, output_file_name)

        # Call the function to process and combine data for each source file
        process_and_combine_data(source_file_path, output_file_path)

print("All data from the files has been processed and saved.")

##get completion date


import re

# Define the date conversion function
def convert_date(input_date):
    # Define a mapping of month names to month numbers
    month_mapping = {
        'January': '01',
        'February': '02',
        'March': '03',
        'April': '04',
        'May': '05',
        'June': '06',
        'July': '07',
        'August': '08',
        'September': '09',
        'October': '10',
        'November': '11',
        'December': '12'
    }

    # Define a regular expression pattern to match the date format
    date_pattern = r'(\w+) (\d{1,2})(?:, )?(\d{4})'

    # Ensure that the input is a string
    if isinstance(input_date, str):
        # Use regular expression to extract parts
        match = re.match(date_pattern, input_date)

        if match:
            # Extract month, day, and year
            month, day, year = match.groups()

            # Get the numeric month value from the mapping
            month = month_mapping.get(month, '00')  

            # Create the formatted date string in 'yyyy-mm-dd' format
            formatted_date = f"{year}-{month}-{day:02}"

            return formatted_date
        else:
            return input_date  # Return the original input for invalid date formats
    else:
        return input_date  # Return the input unchanged for non-string inputs (e.g., datetime objects)

# Apply the date conversion function to the 'completion_date' column after converting it to strings
appended_df['formatted_completion_date'] = appended_df['completion_date'].apply(lambda x: convert_date(x) if isinstance(x, str) else x)


#convert dates to strings
appended_df['formatted_completion_date'] = appended_df['formatted_completion_date'].astype(str)

# Display the DataFrame with the formatted dates
print(appended_df['formatted_completion_date'])


# Define a function to remove the last 9 characters from a cell to remove timestamp manually 
def remove_last_9_characters(cell_value):
    return cell_value[:-9]

# Specify the column to modify
column_name = 'formatted_completion_date'

# Use the apply function to apply the function to the entire column
appended_df['formatted_completion_date'] = appended_df['formatted_completion_date'].apply(remove_last_9_characters)

# Specify the path for the output Excel file
output_file_path = r'\\fs109\ES-Comm\Commercial_Share_Folder\Bright Rebates\BPA Lighting Calculator Extraction Tool - Rosetta\appended_data.xlsx'

# Save the appended DataFrame to a single Excel file
appended_df.to_excel(output_file_path, index=False)

print(f"All dates have been formatted and saved to '{output_file_path}'.")



##get original file path and project id, then join 


import os
import pandas as pd
import re
from openpyxl import load_workbook

# Specify the parent folder containing subfolders with .xls files
parent_folder_path = r'\\fs109\es-comm\Commercial_Share_Folder\Bright Rebates\BPA Lighting Calculator Extraction Tool - Rosetta\ExcelLightCalcInbox\Pre 2019'

# Specify the destination folder path for .xlsx files
destination_folder_path = r'\\fs109\ES-Comm\Commercial_Share_Folder\Bright Rebates\BPA Lighting Calculator Extraction Tool - Rosetta\output_file'

# Create an empty list to store the paths of converted .xlsx files
converted_file_paths = []

# Create empty lists to store parsed components (letter and number) and original .xls paths
letters = []
numbers = []
xls_paths = []  # New list to store the original .xls file paths

# Define the regex pattern for the 'CBR_######' or 'BIZ_######' pattern
pattern = r'(CBR|BIZ)_\s*(\d{5,6})'

# Iterate over subfolders in the parent folder
for subfolder_name in os.listdir(parent_folder_path):
    subfolder_path = os.path.join(parent_folder_path, subfolder_name)

    # Check if the subfolder is a directory
    if os.path.isdir(subfolder_path):
        # Get a list of all files in the subfolder
        file_list = os.listdir(subfolder_path)

        # Loop through each file in the subfolder
        for file_name in file_list:
            # Check if the file is a .xls file
            if file_name.endswith('.xls'):
                xls_path = os.path.join(subfolder_path, file_name)

                # Read specific sheets from the .xls file
                df = pd.read_excel(xls_path, sheet_name=["user_Lighting Details", "UploadData"])
                
                # Concatenate the DataFrames in the dictionary into a single DataFrame
                combined_df = pd.concat(df.values(), ignore_index=True)
                
                # Define the new file name with "loop_" prefix and .xlsx extension
                new_file_name = "loop_" + os.path.splitext(file_name)[0] + ".xlsx"

                # Save the DataFrame as .xlsx with the new file name in the destination folder
                new_xlsx_path = os.path.join(destination_folder_path, new_file_name)
                combined_df.to_excel(new_xlsx_path, index=False)
                
                # Append the path of the converted .xlsx file to the list
                converted_file_paths.append(new_xlsx_path)

                # Use regex to parse the file name and extract the letter and number components
                match = re.search(pattern, file_name)
                if match:
                    letter = match.group(1)
                    number = letter + "_" + match.group(2)  # Combine letter and number
                    letters.append(letter)
                    numbers.append(number)
                else:
                    # Handle cases where the filename does not match the pattern
                    letters.append(None)
                    numbers.append(None)
                
                # Append the original .xls file path to the list
                xls_paths.append(xls_path)

                print(f"Converted '{file_name}' to '{new_file_name}' and saved to destination folder")

# Create a DataFrame containing the list of converted file paths, parsed components, and original .xls paths
data = {
    'code': letters,
    'project_id': numbers,
    'original_path': xls_paths  # Add the original .xls paths to the DataFrame
}

converted_file_paths_df = pd.DataFrame(data)

# Create empty list to store project names
project_names = []

# Iterate over the converted .xlsx paths to extract project names from 'P2'
for xlsx_path in converted_file_paths:
    # Load the Excel file to extract the project name from 'P2' in the new .xlsx file
    wb = load_workbook(xlsx_path, data_only=True)
    sheet = wb.active
    project_name = sheet['P2'].value
    project_names.append(project_name)

# Add the project names to the DataFrame
converted_file_paths_df['project_name'] = project_names

# Display the DataFrame with the paths of converted files, parsed components, and original .xls paths
print("\nList of Converted File Paths, Parsed Components, Project Names, and Original .xls Paths:")
print(converted_file_paths_df)


# Specify the path to save
output_excel_path = r'\\fs109\ES-Comm\Commercial_Share_Folder\Bright Rebates\BPA Lighting Calculator Extraction Tool - Rosetta\converted_files.xlsx'

# Save the DataFrame to an Excel file
converted_file_paths_df.to_excel(output_excel_path, index=False)

print(f"Data saved to '{output_excel_path}'.")


import pandas as pd

# Load the appended_data.xlsx file
appended_data_df = pd.read_excel('appended_data.xlsx')

# Merge the two DataFrames based on the 'project_name' column
merged_df = pd.merge(appended_data_df, converted_file_paths_df, on='project_name', how='left')


##change column names

# Rename columns in df_existing
column_rename_dict = {
    'est_annual_energy_savings':'est_annual_project_total_energy_savings',
    'est_total_proj_install_costs':'est_total_project_cost',
    'projcost':'final_total_project_cost',
    'busbr_sav':'est_total_busbar_savings',
    'wattred':'wattage_reduction',
    'fedblg':'federal_building_flag',
    'bpa_max_wtp':'BPA_willingness_to_pay',
    'incentives':'incentive_notes',
    'notes':'measure_notes',
    'electric_interaction':'HVAC_interaction_factor',
    'precond_decom_and_equip_kwh_savings_with_hvac':'Retrofit_and_Decom_Measure_Savings_HVAC_Ad',
    'precond_equip_and_controls_kwh_savings_with_hvac_busbar': 'Retrofit_and_Decom_Measure_Savings_HVAC_Busbar_Adj',
    'precond_total_measure_kwh':'baseline_total_fixture_kWh',
    'precond_total_measure_kwh_with_hvac':'baseline_total_fixture_kWh_HVAC_Adj',
    'proposed_total_measure_kwh_with_hvac': 'proposed_total_fixture_kWh_HVAC_Adj'
    }


merged_df.rename(columns=column_rename_dict, inplace=True)


##QA TA contact name

name_mapping = {
    'Andy D':'Andy Deweyert',
    'Bryan Gray':'Bryan Grey',
    'Christina Henrickson':'Christina Henricksen',
    'ED BRAY':'Ed Bray',
    'self install': '',
    'Self Install': '', 
    'Jeffery Gascoyne':'Jeff Gascoyne',
    'Jerome Guissler': 'Jerome Geissler',
    'Thomas Harwood':'Tom Harwood',
    'Frederik Wouda Kuipers':'Fred Wouda Kuipers',
    'self':'',
    'Self install':'',
    'JEFF GASCOYNE SR':'Jeff Gascoyne',
    'Mark Niemanmarkn@mckinstry.com':'Mark Nieman'  
    
}

# Use the replace method to replace values in the 'ta_organization' column
merged_df['ta_contact_name'] = merged_df['ta_contact_name'].replace(name_mapping)

merged_df['ta_contact_name'] = merged_df['ta_contact_name'].str.upper()


# QA TA Org name
rename_mapping = {
    'BidEnergy': 'Bid Energy',
    'Lumenal Lighting': 'Lumenal Lighting LLC',
    'North Coast': 'North Coast Electric',
    'North West Edison': 'Northwest Edison',
    'Platt Electric': 'Platt Electric Supply',
    'RAINIER LIGHTING AND ELECTRICAL SUPPLY': 'Rainier Lighting & Electric Supply',
    'Seahurst Electic': 'Seahurst Electric',
    'self ': 'Self Install',
    'self':'Self Install',
    'Sellf Install': 'Self Install',
    'self install':'Self Install',
    'Self install': 'Self Install',
    'Stoneway Electrical Supply':'Stoneway Electric Supply',
    'UNited Lamp ': 'United Lamp Supply',
    'Wesco': 'WESCO Energy Solutions',
    'Resound Energy': 'Resound Energy LLC',
    'NW Edison': 'Northwest Edison',
    'Source One Solutions': 'SourceOne Solutions',
    'McKinstry': 'McKinstry Essention, LLC',
    'Allied Electric': 'Allied Electric Corporation',
    'Amaya': 'Amaya Electric',
    'Ameya Eletric':'Amaya Eletric',
    'Ault Electtric':'Ault Electric',
    'Beyond Basic':'Beyond Basic Electric, In.',
    'Capitol Light':'Capital Lighting',
    'Creative Lighting':'Creative Lighting Solutions, Inc.',
    'Danard Electric Inc.':'Danard Electric',
    'DC Engineering Inc.':'DC Engineering',
    'Eco Engineering.com':'Eco Engineering Inc.',
    'Energy Management Collaborative':'Energy Management Collaborative LLC',
    'Energy Management Collaborative llc':'Energy Management Collaborative LLC',
    'Energy Retrofit co':'Energy Retrofit Co.',
    'Energy Retrofit Co':'Energy Retrofit Co.',
    'Graybar': 'Graybar Electric',
    'Green Lighting LLC':'Green Lighting, LLC',
    'Leidos Engineering LLC.':'Leidos Engineering, LLC',
    'Leidos Engineering LLC':'Leidos Engineering, LLC',
    'Lighitng & Power': 'Lighting & Power, Inc.',
    'Lighting & Power Inc.': 'Lighting & Power, Inc',
    'Lighting&Power, Inc':'Lighting & Power, Inc',
    'Lumenal': 'Lumenal Lighting LLC',
    'Lumenal Lighing formerly known as Light Doctor': 'Lumenal Lighting LLC',
    'Lumenal Lighting':'Lumenal Lighting LLC',
    'Lumenal Lighting formely known as Light Doctor':'Lumenal Lighting LLC',
    'Lumenal Lighting formerly known as Light Doctor':'Lumenal Lighting LLC',
    'Lumenal Lighting formerly known Light Doctor':'Lumenal Lighting LLC',
    'Lumenal Lighting, LLC.':'Lumenal Lighting LLC',
    'Lumenal Ligitng formerly known as Light Doctor':'Lumenal Lighting LLC',
    'Lumenal Ligthing Formely known as Light Doctor':'Lumenal Lighting LLC',
    'McKinstry Electric':'McKinstry Essention, LLC',
    'Narrows Heating & AC, Inc.':'Narrows Heating & Air Conditioning',
    'NW Lighting Solutions':'NW LIGHTING SOLUTIONS, LLC',
    'NW LIGHTING SOLUTIONS,LLC':'NW LIGHTING SOLUTIONS, LLC',
    'Olsen Electric':'Olsen Electric Inc.',
    'Pacific Energy Concepts':'Pacific Energy Concepts, LLC',
    'Pacific Lamp & Supply':'Pacific Lamp & Supply Co',
    'PlanLED':'PlanLED Inc.',
    'Platt':'Platt Electric Supply',
    'Resound Energy Services':'Resound Energy Services, LLC',
    'Resound Energy LLC':'Resound Energy Services, LLC',
    'Resound Eneryg Services':'Resound Energy Services, LLC',
    'Seahurst Electgric':'Seahurst Electric',
    'Stoneay Electric Supply':'Stoneway Electric Supply',
    'STONEWAY ELECTRIC':'Stoneway Electric Supply',
    'Stoneway Electric Suppley':'Stoneway Electric Supply',
    'StonewayElectric Supply':'Stoneway Electric Supply',
    'Sylvania Lighting Solutions':'Sylvania Lighting Services',
    'Tacoma Electric':'TACOMA ELECTRIC SUPPLY, INC',
    'TACOMA ELECTRIC SUPPLY':'TACOMA ELECTRIC SUPPLY, INC',
    'Tacoma Electric Supply':'TACOMA ELECTRIC SUPPLY, INC',
    'United Lamp':'United Lamp Supply'
    
}

# Use the replace method to replace values in the 'ta_organization' column
merged_df['ta_organization'] = merged_df['ta_organization'].replace(rename_mapping)

#everything to upper case

merged_df['ta_organization'] = merged_df['ta_organization'].str.upper() 

# Display the DataFrame with the updated 'ta_organization' column
print(merged_df['ta_organization'])

print(f"Merged data saved.")

# Specify the columns to move and their new positions
columns_to_move = ['mailing_full_address', 'project/site_full_address']
new_positions = [1, 2]

# Get the list of all columns in the DataFrame
columns = merged_df.columns.tolist()

# Iterate over the columns to move and their new positions
for col, pos in zip(columns_to_move, new_positions):
    columns.remove(col)  # Remove the column from its current position
    columns.insert(pos, col)  # Insert the column at the new position

# Create a new DataFrame with the columns reordered
merged_df = merged_df[columns]

# Display the modified DataFrame
print(merged_df)

# Drop one or more columns
columns_to_drop = ['tpu_contact_phone', 'tpu_email', 'ta_contact_phone', 'est_energy_savings_(%)',
                  'totcred', 'status', 'existing', 'proposed', 'blank_column', 'kw_for_table', 'proposed_lamp_model_number',
                  'proposed_ballast_model_number', 'no_controls', 'number_of_errors','existing_mml_row', 'proposed_mml_row',
                   'nonstandard_incentive_rates', 'completion_date' 
                  ]  # List of columns to drop

# Use the drop method to drop the specified columns
merged_df = merged_df.drop(columns=columns_to_drop)


# Specify the path to save the merged DataFrame as a new Excel file
output_merged_excel_path = r'\\fs109\ES-Comm\Commercial_Share_Folder\Bright Rebates\BPA Lighting Calculator Extraction Tool - Rosetta\lc_final_draft_v0.xlsx'

# Save the merged DataFrame to a new Excel file
merged_df.to_excel(output_merged_excel_path, index=False)

print(f"Merged data saved to '{output_merged_excel_path}'. This is the final dataset")